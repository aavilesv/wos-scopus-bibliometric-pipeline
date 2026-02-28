# ============================================================
# reporting.py
#   - Guarda outputs principales (CSV)
#   - Genera y guarda reportes (Excel) de métricas/tablas
#   - Genera gráficos: mostrar + guardar PNG
# ============================================================
from __future__ import annotations

from pathlib import Path
from typing import Set, Optional, Dict, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from ui_messages import info, warn


# ------------------------------------------------------------
# CSV con tildes OK en Excel (UTF-8 con BOM)
# ------------------------------------------------------------
def _save_csv_utf8sig(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


# ------------------------------------------------------------
# Combinar Author Keywords + Index Keywords
# ------------------------------------------------------------
def _create_bothkeywords_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea columna 'bothkeywords' combinando 'Author Keywords' e 'Index Keywords'.
    Si alguna está vacía, usa solo la disponible.
    """
    df = df.copy()
    
    author_kw_col = "Author Keywords"
    index_kw_col = "Index Keywords"
    
    # Asegurarnos de que existan las columnas
    if author_kw_col not in df.columns:
        df[author_kw_col] = ""
    if index_kw_col not in df.columns:
        df[index_kw_col] = ""
    
    # Convertir a string y rellenar NaN
    df[author_kw_col] = df[author_kw_col].fillna("").astype(str).str.strip()
    df[index_kw_col] = df[index_kw_col].fillna("").astype(str).str.strip()
    
    # Combinar y eliminar duplicados
    def combine_keywords(row):
        author = row[author_kw_col]
        index = row[index_kw_col]
        
        # Recolectar todas las palabras clave
        all_keywords = []
        
        if author:
            # Dividir por ";" y limpiar espacios
            author_kws = [kw.strip() for kw in author.split(";") if kw.strip()]
            all_keywords.extend(author_kws)
        
        if index:
            # Dividir por ";" y limpiar espacios
            index_kws = [kw.strip() for kw in index.split(";") if kw.strip()]
            all_keywords.extend(index_kws)
        
        if not all_keywords:
            return ""
        
        # Eliminar duplicados manteniendo el orden (case-insensitive comparison)
        seen_lower = set()
        unique_keywords = []
        for kw in all_keywords:
            kw_lower = kw.lower()
            if kw_lower not in seen_lower:
                seen_lower.add(kw_lower)
                unique_keywords.append(kw)
        
        # Unir con "; "
        return "; ".join(unique_keywords)
    
    df["bothkeywords"] = df.apply(combine_keywords, axis=1)
    
    return df


# ------------------------------------------------------------
# Excel reporte (solo tablas/resúmenes)
# ------------------------------------------------------------
def _save_report_excel(sheets: Dict[str, pd.DataFrame], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=name[:31], index=False)

    # Formato simple (si falla, igual queda el xlsx)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_align

            max_cap = 60
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 80)), start=1):
                lengths = []
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    lengths.append(len(val))
                    if cell.row > 1:
                        cell.alignment = cell_align

                width = min((max(lengths) + 2) if lengths else 12, max_cap)
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, width)

        wb.save(path)
    except Exception as e:
        warn("Formato Excel", f"No se pudo aplicar formato avanzado al reporte.\nDetalle: {e}")


# ------------------------------------------------------------
# 1) Guardado de outputs principales
# ------------------------------------------------------------
def save_outputs(
    combined_df: pd.DataFrame,
    duplicated_titles: Set[str],
    results_dir: Path
) -> Tuple[Path, Path]:
    """
    Guarda:
      - datawos_scopus.csv (con columna bothkeywords)
      - datawos_scopus_repeatedstitles.csv
    Retorna rutas guardadas.
    """
    results_dir.mkdir(parents=True, exist_ok=True)

    out_main = results_dir / "datawos_scopus.csv"
    out_dups = results_dir / "datawos_scopus_repeatedstitles.csv"

    if combined_df is None:
        combined_df = pd.DataFrame()
    else:
        # Crear columna bothkeywords: combina Author Keywords + Index Keywords
        combined_df = _create_bothkeywords_column(combined_df)

    dups_df = pd.DataFrame(sorted(list(duplicated_titles)), columns=["Título Repetido"])

    _save_csv_utf8sig(combined_df, out_main)
    _save_csv_utf8sig(dups_df, out_dups)

    info(
        "Archivos guardados",
        "Se guardaron los outputs principales:\n\n"
        f"- {out_main}\n"
        f"- {out_dups}"
    )

    return out_main, out_dups

def compute_quality_metrics(df):
    if df is None or df.empty:
        return 0, 0, 0, 0, 0

    total = len(df)

    valid_dois = (
        df.get("DOI", pd.Series())
        .fillna("")
        .astype(str)
        .str.strip() != ""
    ).sum()

    valid_titles = (
        df.get("Title", pd.Series())
        .fillna("")
        .astype(str)
        .str.strip() != ""
    ).sum()

    valid_authors = (
        df.get("Authors", pd.Series())
        .fillna("")
        .astype(str)
        .str.strip() != ""
    ).sum()

    valid_abstracts = (
        df.get("Abstract", pd.Series())
        .fillna("")
        .astype(str)
        .str.strip() != ""
    ).sum()

    return (
        total,
        valid_dois,
        valid_titles,
        valid_authors,
        valid_abstracts
    )

# ------------------------------------------------------------
# 2) Construcción de tablas de reporte (lo que antes era print)
# ------------------------------------------------------------
def build_report_tables(
    *,
    original_scopus_count: int,
    original_wos_count: int,
    scopus_df: Optional[pd.DataFrame],
    wos_df: Optional[pd.DataFrame],
    wos_non_repeated: Optional[pd.DataFrame],
    df_wos_renombrado: Optional[pd.DataFrame],
    combined_df: Optional[pd.DataFrame],
    duplicated_titles: Set[str],
    year_start: int,
    year_end: int,
    scopus_stats: Optional[Dict] = None,
    wos_stats: Optional[Dict] = None,
    removed_post_merge: int = 0
) -> Dict[str, pd.DataFrame]:
    """
    Devuelve un dict {sheet_name: df} para exportar a Excel.
    Replica la lógica de tus prints, pero en tablas.
    """

    # --- Counts for Funnel ---
    
    # 1. Raw Loaded (Identification)
    raw_scopus = int(original_scopus_count)
    raw_wos = int(original_wos_count)
    total_loaded = raw_scopus + raw_wos
    
    scopus_int_dups = scopus_stats.get('internal_duplicates', 0) if scopus_stats else 0
    wos_int_dups = wos_stats.get('internal_duplicates', 0) if wos_stats else 0
    
    # 2. After Internal Dedup
    after_int_scopus = raw_scopus - scopus_int_dups
    after_int_wos = raw_wos - wos_int_dups
    
    # 3. Quality Removed
    sc_qty_rem = 0
    if scopus_stats:
        sc_qty_rem = scopus_stats.get('removed_doi',0) + scopus_stats.get('removed_title',0) + \
                     scopus_stats.get('removed_authors',0) + scopus_stats.get('removed_abstract',0)
                     
    wo_qty_rem = 0
    if wos_stats:
        wo_qty_rem = wos_stats.get('removed_doi',0) + wos_stats.get('removed_title',0) + \
                     wos_stats.get('removed_authors',0) + wos_stats.get('removed_abstract',0)
                     
    after_qual_scopus = after_int_scopus - sc_qty_rem
    after_qual_wos = after_int_wos - wo_qty_rem
    after_qual_total = after_qual_scopus + after_qual_wos
    
    # 4. Cross Dedup (assuming usually cross duplicates are removed from WoS)
    cross_dups = len(duplicated_titles)
    
    # 5. Final Counts
    if combined_df is not None and not combined_df.empty:
        total_final = len(combined_df)
        sources_series = combined_df.get("Source", pd.Series(dtype=str)).astype(str).str.lower()
        final_scopus = int((sources_series == "scopus").sum())
        final_wos = total_final - final_scopus  # The rest are WoS
    else:
        total_final = 0
        final_scopus = 0
        final_wos = 0

    # Removed at the end due to cross deduplication + post-merge cleanup
    # We attribute unknown removals proportionately or based on theoretical flows
    total_removed_scopus = raw_scopus - final_scopus
    total_removed_wos = raw_wos - final_wos

    # Proportions
    raw_scopus_pct = round((raw_scopus / total_loaded * 100), 1) if total_loaded else 0
    raw_wos_pct = round((raw_wos / total_loaded * 100), 1) if total_loaded else 0
    
    final_scopus_pct = round((final_scopus / total_final * 100), 1) if total_final else 0
    final_wos_pct = round((final_wos / total_final * 100), 1) if total_final else 0

    removed_scopus_pct = round((total_removed_scopus / raw_scopus * 100), 1) if raw_scopus else 0
    removed_wos_pct = round((total_removed_wos / raw_wos * 100), 1) if raw_wos else 0

    # Dataset before the final merge
    if scopus_df is not None and wos_df is not None:
        before_df = pd.concat([scopus_df, wos_df], ignore_index=True)
    elif scopus_df is not None:
        before_df = scopus_df
    elif wos_df is not None:
        before_df = wos_df
    else:
        before_df = None

    before_total, before_doi, before_title, before_auth, before_abs = compute_quality_metrics(before_df)

    valid_dois = valid_titles = valid_authors = valid_abstracts = 0
    doi_pct = title_pct = author_pct = abstract_pct = 0
    
    if combined_df is not None and not combined_df.empty:
        valid_dois = (combined_df.get("DOI", pd.Series()).fillna("").astype(str).str.strip() != "").sum()
        valid_titles = (combined_df.get("Title", pd.Series()).fillna("").astype(str).str.strip() != "").sum()
        valid_authors = (combined_df.get("Authors", pd.Series()).fillna("").astype(str).str.strip() != "").sum()
        valid_abstracts = (combined_df.get("Abstract", pd.Series()).fillna("").astype(str).str.strip() != "").sum()
        
        doi_pct = round(valid_dois / total_final * 100, 1) if total_final else 0
        title_pct = round(valid_titles / total_final * 100, 1) if total_final else 0
        author_pct = round(valid_authors / total_final * 100, 1) if total_final else 0
        abstract_pct = round(valid_abstracts / total_final * 100, 1) if total_final else 0

    stats_summary = pd.DataFrame(
        [
            ("1. RAW DATA", ""),
            ("Loaded papers (Total)", total_loaded),
            ("Loaded WoS", raw_wos),
            ("Loaded WoS (%)", raw_wos_pct),
            ("Loaded Scopus", raw_scopus),
            ("Loaded Scopus (%)", raw_scopus_pct),
            ("", ""),
            ("2. DEDUPLICATION & FILTERING TOTALS", ""),
            ("Internal Duplicates Removed (Scopus)", scopus_int_dups),
            ("Internal Duplicates Removed (WoS)", wos_int_dups),
            ("Quality Failed Records Removed (Scopus)", sc_qty_rem),
            ("Quality Failed Records Removed (WoS)", wo_qty_rem),
            ("Cross-Database Duplicates Removed", cross_dups),
            ("Post-Merge/Year Filtering Removed", removed_post_merge),
            ("", ""),
            ("3. FINAL DATASET", ""),
            ("Total Removed (Scopus)", total_removed_scopus),
            ("Total Removed (Scopus %)", removed_scopus_pct),
            ("Total Removed (WoS)", total_removed_wos),
            ("Total Removed (WoS %)", removed_wos_pct),
            ("Total Resulting Papers", total_final),
            ("Final WoS", final_wos),
            ("Final WoS (%)", final_wos_pct),
            ("Final Scopus", final_scopus),
            ("Final Scopus (%)", final_scopus_pct),
            ("", ""),
            ("QUALITY CHECK (Before Filters)", ""),
            ("Total records before quality filters", before_total),
            ("Records with DOI", before_doi),
            ("Records with DOI (%)", round(before_doi / before_total * 100, 1) if before_total else 0),
            ("Records with Title", before_title),
            ("Records with Title (%)", round(before_title / before_total * 100, 1) if before_total else 0),
            ("Records with Authors", before_auth),
            ("Records with Authors (%)", round(before_auth / before_total * 100, 1) if before_total else 0),
            ("Records with Abstract", before_abs),
            ("Records with Abstract (%)", round(before_abs / before_total * 100, 1) if before_total else 0),
            ("", ""),
            ("QUALITY CHECK (Final Dataset)", ""),
            ("Total records after quality filters", total_final),
            ("Records with DOI", valid_dois),
            ("Records with DOI (%)", doi_pct),
            ("Records with Title", valid_titles),
            ("Records with Title (%)", title_pct),
            ("Records with Authors", valid_authors),
            ("Records with Authors (%)", author_pct),
            ("Records with Abstract", valid_abstracts),
            ("Records with Abstract (%)", abstract_pct),
        ],
        columns=["Metric", "Value"]
    )

    # --- Distribución kept/removed por fuente (para gráfico y tabla) ---
    dedup_distribution = pd.DataFrame(
        {
            "Source": ["WoS", "Scopus"],
            "Kept": [final_wos, final_scopus],
            "Removed": [total_removed_wos, total_removed_scopus],
        }
    )
    dedup_distribution["Total"] = dedup_distribution["Kept"] + dedup_distribution["Removed"]
    dedup_distribution["Kept (%)"] = np.where(
        dedup_distribution["Total"] > 0,
        (dedup_distribution["Kept"] / dedup_distribution["Total"] * 100).round(1),
        0.0
    )
    dedup_distribution["Removed (%)"] = np.where(
        dedup_distribution["Total"] > 0,
        (dedup_distribution["Removed"] / dedup_distribution["Total"] * 100).round(1),
        0.0
    )

    # --- Tablas RAW por año (antes de dedup) ---
    raw_counts = pd.DataFrame()
    raw_citations = pd.DataFrame()

    try:
        if scopus_df is not None and "Year" in scopus_df.columns:
            sc_year = pd.to_numeric(scopus_df["Year"], errors="coerce")
            mask_sc = sc_year.between(year_start, year_end)
            scopus_yearly = scopus_df.loc[mask_sc].assign(Year=sc_year[mask_sc]).groupby("Year").size()
        else:
            scopus_yearly = pd.Series(dtype=int)

        # wos_df tiene Publication Year en tu export original
        if wos_df is not None and "Publication Year" in wos_df.columns:
            wo_year = pd.to_numeric(wos_df["Publication Year"], errors="coerce")
            mask_wo = wo_year.between(year_start, year_end)
            wos_yearly = wos_df.loc[mask_wo].assign(Year=wo_year[mask_wo]).groupby("Year").size()
        else:
            wos_yearly = pd.Series(dtype=int)

        raw_counts = pd.DataFrame({"WoS": wos_yearly, "Scopus": scopus_yearly}).fillna(0).astype(int)
        raw_counts = raw_counts.reset_index().rename(columns={"index": "Year"})
        raw_counts["Total Articles Raw"] = raw_counts["WoS"] + raw_counts["Scopus"]

        # Citas RAW
        if scopus_df is not None and "Cited by" in scopus_df.columns and "Year" in scopus_df.columns:
            sc_year = pd.to_numeric(scopus_df["Year"], errors="coerce")
            mask_sc = sc_year.between(year_start, year_end)
            sc_cites = pd.to_numeric(scopus_df.loc[mask_sc, "Cited by"], errors="coerce").fillna(0)
            scopus_cites = pd.DataFrame({"Year": sc_year[mask_sc], "Cited by": sc_cites}).groupby("Year")["Cited by"].sum()
        else:
            scopus_cites = pd.Series(dtype=int)

        if wos_df is not None and "Cited Reference Count" in wos_df.columns and "Publication Year" in wos_df.columns:
            wo_year = pd.to_numeric(wos_df["Publication Year"], errors="coerce")
            mask_wo = wo_year.between(year_start, year_end)
            wo_cites = pd.to_numeric(wos_df.loc[mask_wo, "Cited Reference Count"], errors="coerce").fillna(0)
            wos_cites = pd.DataFrame({"Year": wo_year[mask_wo], "Cited Reference Count": wo_cites}).groupby("Year")["Cited Reference Count"].sum()
        else:
            wos_cites = pd.Series(dtype=int)

        raw_citations = pd.DataFrame({"WoS Citations": wos_cites, "Scopus Citations": scopus_cites}).fillna(0).astype(int)
        raw_citations = raw_citations.reset_index().rename(columns={"index": "Year"})
        raw_citations["Total Citations Raw"] = raw_citations["WoS Citations"] + raw_citations["Scopus Citations"]

    except Exception as e:
        warn("Reporte RAW por año", f"No se pudieron construir tablas RAW por año.\nDetalle: {e}")

    # --- Tabla Document Type por año (si existe) ---
    doc_types_by_year = pd.DataFrame()
    try:
        if combined_df is not None and {"Year", "Document Type"}.issubset(set(combined_df.columns)):
            df = combined_df[["Year", "Document Type"]].copy()
            df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
            df = df.dropna(subset=["Year"])
            df = df[df["Year"].between(year_start, year_end)]

            df["Document Type"] = df["Document Type"].astype(str).str.split(";")
            df["Document Type"] = df["Document Type"].apply(
                lambda x: [i.strip() for i in x if i.strip()] if isinstance(x, list) else []
            )
            exploded = df.explode("Document Type")
            exploded = exploded[exploded["Document Type"].astype(str).str.strip() != ""]

            pivot = exploded.groupby(["Year", "Document Type"]).size().unstack(fill_value=0)
            doc_types_by_year = pivot.reset_index()
    except Exception as e:
        warn("Reporte Document Type", f"No se pudo construir tabla de Document Type por año.\nDetalle: {e}")

    # --- PRISMA Flow ---
    prisma_steps = []
    
    # 1. Identification
    sc_internal_dups = scopus_stats.get("internal_duplicates", 0) if scopus_stats else 0
    wo_internal_dups = wos_stats.get("internal_duplicates", 0) if wos_stats else 0
    
    prisma_steps.append(("1. Records identified from Scopus", original_scopus_count))
    prisma_steps.append(("1. Records identified from WoS", original_wos_count))
    prisma_steps.append(("Records identified (Total)", total_loaded))
    prisma_steps.append(("", ""))
    
    # 2. Screening - Internal Duplicates
    prisma_steps.append(("2. Internal duplicates removed (Scopus)", sc_internal_dups))
    prisma_steps.append(("2. Internal duplicates removed (WoS)", wo_internal_dups))
    prisma_steps.append(("Records after internal deduplication", total_loaded - sc_internal_dups - wo_internal_dups))
    prisma_steps.append(("", ""))
    
    # 3. Screening - Quality Checks (Scopus)
    sc_doi = scopus_stats.get("removed_doi", 0) if scopus_stats else 0
    sc_title = scopus_stats.get("removed_title", 0) if scopus_stats else 0
    sc_author = scopus_stats.get("removed_authors", 0) if scopus_stats else 0
    sc_abstract = scopus_stats.get("removed_abstract", 0) if scopus_stats else 0
    
    prisma_steps.append(("3. Missing/Invalid DOIs removed (Scopus)", sc_doi))
    prisma_steps.append(("3. Missing Titles removed (Scopus)", sc_title))
    prisma_steps.append(("3. Missing Authors removed (Scopus)", sc_author))
    prisma_steps.append(("3. Missing Abstracts removed (Scopus)", sc_abstract))
    
    # 4. Screening - Quality Checks (WoS)
    wo_doi = wos_stats.get("removed_doi", 0) if wos_stats else 0
    wo_title = wos_stats.get("removed_title", 0) if wos_stats else 0
    wo_author = wos_stats.get("removed_authors", 0) if wos_stats else 0
    wo_abstract = wos_stats.get("removed_abstract", 0) if wos_stats else 0
    
    prisma_steps.append(("4. Missing/Invalid DOIs removed (WoS)", wo_doi))
    prisma_steps.append(("4. Missing Titles removed (WoS)", wo_title))
    prisma_steps.append(("4. Missing Authors removed (WoS)", wo_author))
    prisma_steps.append(("4. Missing Abstracts removed (WoS)", wo_abstract))
    
    records_after_quality = (scopus_stats.get("final_count", 0) if scopus_stats else scopus_unique_count) + \
                            (wos_stats.get("final_count", 0) if wos_stats else wos_non_rep_count + removed_wos) # Approximation if stats missing
                            
    prisma_steps.append(("Records screened for quality", records_after_quality))
    prisma_steps.append(("", ""))
    
    prisma_steps.append(("5. Cross-database duplicates removed", cross_dups))
    prisma_steps.append(("5. Post-Merge / Year filtering removed", removed_post_merge))
    prisma_steps.append(("Total records included for synthesis", total_final))
    
    prisma_flow = pd.DataFrame(prisma_steps, columns=["PRISMA Step", "Count"])

    # --- Methodology ---
    methodology_steps = [
        ("Step", "Description", "Python/Library Details"),
        ("1. Data Loading", "Load CSVs from Scopus and Excel from WoS.", "pandas.read_csv, pandas.read_excel"),
        ("2. Internal Deduplication", "Remove duplicate rows within the same database based on Title.", "pandas.drop_duplicates(subset=['processed_title'])"),
        ("3. Quality: DOI Cleaning", "Extract valid DOIs using regular expressions and remove empty.", "re.search(r'(10\\.\\d{4,9}/[-._;()/:a-z0-9]+)'), df[df['DOI'] != '']"),
        ("4. Quality: Missing Titles", "Remove records where the title is completely empty.", "df[df['Title'].fillna('').str.strip() != '']"),
        ("5. Quality: Missing Authors", "Remove records where authors are missing.", "df[df['Authors'].fillna('').str.strip() != '']"),
        ("6. Quality: Missing Abstracts", "Remove records where abstract is missing.", "df[df['Abstract'].fillna('').str.strip() != '']"),
        ("7. Cross-Deduplication", "Fuzzy matching to find overlapping titles between Scopus and WoS.", "thefuzz.fuzz.token_set_ratio > config.FUZZY_THRESHOLD"),
        ("8. Merging", "Normalize WoS columns to Scopus schema and concatenate.", "pandas.concat([scopus, wos_normalized])"),
        ("9. Enriching", "Add SCImago Quartiles and Metrics based on ISSN/Journal Title.", "pandas.merge(combined_df, scimago_df)")
    ]
    
    methodology_df = pd.DataFrame(methodology_steps[1:], columns=methodology_steps[0])

    # --- Tabla de Métricas de Calidad Final ---
    quality_summary = pd.DataFrame()
    try:
        if combined_df is not None and not combined_df.empty:
            total_final = len(combined_df)
            
            # DOIs
            valid_dois = (combined_df.get("DOI", pd.Series()).fillna("").astype(str) != "").sum()
            blank_dois = total_final - valid_dois
            
            # Years
            years_numeric = pd.to_numeric(combined_df.get("Year", pd.Series()), errors='coerce')
            valid_years = years_numeric.notna().sum()
            blank_years = years_numeric.isna().sum()
            
            # Titles
            valid_titles = (combined_df.get("Title", pd.Series()).fillna("").astype(str).str.strip() != "").sum()
            blank_titles = total_final - valid_titles
            
            # bothkeywords (si existe)
            has_keywords = 0
            if "bothkeywords" in combined_df.columns:
                has_keywords = (combined_df["bothkeywords"].fillna("").astype(str).str.strip() != "").sum()
            
            quality_summary = pd.DataFrame([
                ("Total Records (Final)", total_final),
                ("", ""),
                ("DOIs - Valid", valid_dois),
                ("DOIs - Blank", blank_dois),
                ("DOIs - Valid (%)", round(valid_dois/total_final*100, 1) if total_final else 0),
                ("", ""),
                ("Years - Valid", valid_years),
                ("Years - Missing", blank_years),
                ("Years - Valid (%)", round(valid_years/total_final*100, 1) if total_final else 0),
                ("", ""),
                ("Titles - Valid", valid_titles),
                ("Titles - Blank", blank_titles),
                ("Titles - Valid (%)", round(valid_titles/total_final*100, 1) if total_final else 0),
                ("", ""),
                ("Keywords - With Data", has_keywords),
                ("Keywords - Blank", total_final - has_keywords),
                ("Keywords - With Data (%)", round(has_keywords/total_final*100, 1) if total_final else 0),
            ], columns=["Metric", "Value"])
    except Exception as e:
        warn("Reporte Calidad", f"No se pudo construir tabla de calidad final.\nDetalle: {e}")

    return {
        "stats_summary": stats_summary,
        "PRISMA_Flow": prisma_flow,
        "Methodology": methodology_df,
        "quality_summary": quality_summary,
        "dedup_distribution": dedup_distribution,
        "raw_counts_by_year": raw_counts,
        "raw_citations_by_year": raw_citations,
        "doc_types_by_year": doc_types_by_year,
    }


def save_report_excel(report_tables: Dict[str, pd.DataFrame], results_dir: Path) -> Path:
    """
    Guarda un solo Excel de reporte con varias hojas.
    """
    results_dir.mkdir(parents=True, exist_ok=True)
    out = results_dir / "report.xlsx"
    _save_report_excel(report_tables, out)

    info("Reporte Excel", f"Se guardó el reporte en Excel:\n{out}")
    return out


# ------------------------------------------------------------
# 3) Gráfico distribución kept/removed (mostrar + guardar)
# ------------------------------------------------------------
def plot_distribution(
    final_wos: int,
    final_scopus: int,
    removed_wos: int,
    removed_scopus: int,
    results_dir: Optional[Path] = None,
    filename: str = "distribution_post_dedup.png",
    show: bool = True,
    dpi: int = 300,
) -> Optional[Path]:
    sources = ["WoS", "Scopus"]
    kept = [final_wos, final_scopus]
    removed = [removed_wos, removed_scopus]

    totals = np.array(kept) + np.array(removed)
    order = np.argsort(totals)[::-1]

    sources = [sources[i] for i in order]
    kept = [kept[i] for i in order]
    removed = [removed[i] for i in order]
    totals = (np.array(kept) + np.array(removed)).astype(float)

    pct_kept = np.where(totals > 0, np.array(kept) / totals * 100, 0.0)
    pct_removed = np.where(totals > 0, np.array(removed) / totals * 100, 0.0)

    fig, ax = plt.subplots(figsize=(8, 4))
    bars_kept = ax.barh(sources, kept, label="Kept")
    bars_removed = ax.barh(sources, removed, left=kept, label="Removed")

    # etiquetas (como tu script)
    for i, (b1, b2) in enumerate(zip(bars_kept, bars_removed)):
        w1 = b1.get_width()
        ax.text(
            w1 / 2 if w1 else 0,
            b1.get_y() + b1.get_height() / 2,
            f"{kept[i]}\n({pct_kept[i]:.1f}%)",
            va="center",
            ha="center" if w1 else "left",
        )

        w2 = b2.get_width()
        if w2 > 0:
            ax.text(
                kept[i] + w2 / 2,
                b2.get_y() + b2.get_height() / 2,
                f"{removed[i]}\n({pct_removed[i]:.1f}%)",
                va="center",
                ha="center",
            )

    ax.invert_yaxis()
    ax.set_title(
        "Post-deduplication Distribution of Bibliometric Records\nfrom Scopus and Web of Science",
        weight="bold",
        pad=12,
    )
    ax.set_xlabel("Number of Articles")
    ax.legend(loc="lower right")
    ax.grid(axis="x", linestyle="--", alpha=0.5)
    plt.tight_layout()

    saved_path = None
    if results_dir is not None:
        results_dir.mkdir(parents=True, exist_ok=True)
        saved_path = results_dir / filename
        fig.savefig(saved_path, dpi=dpi, bbox_inches="tight")
        info("Gráfico guardado", f"El gráfico se guardó en:\n{saved_path}")

    if show:
        plt.show()
    else:
        plt.close(fig)

    return saved_path


# ------------------------------------------------------------
# 4) Gráficos RAW por año (mostrar + guardar)
# ------------------------------------------------------------
def plot_raw_trends(
    raw_counts_by_year: pd.DataFrame,
    raw_citations_by_year: pd.DataFrame,
    results_dir: Optional[Path] = None,
    show: bool = True,
    dpi: int = 300,
) -> Tuple[Optional[Path], Optional[Path]]:
    """
    Guarda (opcional) y muestra:
      - raw_articles_by_year.png
      - raw_citations_by_year.png
    """
    saved_counts = None
    saved_cites = None

    # --- Artículos por año ---
    if raw_counts_by_year is not None and not raw_counts_by_year.empty and "Year" in raw_counts_by_year.columns:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(raw_counts_by_year["Year"], raw_counts_by_year.get("WoS", 0), marker="o", label="WoS Articles")
        ax.plot(raw_counts_by_year["Year"], raw_counts_by_year.get("Scopus", 0), marker="s", label="Scopus Articles")
        ax.plot(raw_counts_by_year["Year"], raw_counts_by_year.get("Total Articles Raw", 0), marker="^", label="Total Articles")
        ax.set_title("Annual evolution of articles (RAW data before deduplication)", weight="bold", pad=12)
        ax.set_xlabel("Year")
        ax.set_ylabel("Number of articles")
        ax.legend(loc="upper left")
        ax.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()

        if results_dir is not None:
            results_dir.mkdir(parents=True, exist_ok=True)
            saved_counts = results_dir / "raw_articles_by_year.png"
            fig.savefig(saved_counts, dpi=dpi, bbox_inches="tight")
            info("Gráfico guardado", f"RAW artículos por año:\n{saved_counts}")

        if show:
            plt.show()
        else:
            plt.close(fig)

    # --- Citas por año ---
    if raw_citations_by_year is not None and not raw_citations_by_year.empty and "Year" in raw_citations_by_year.columns:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(raw_citations_by_year["Year"], raw_citations_by_year.get("WoS Citations", 0), marker="o", label="WoS Citations")
        ax.plot(raw_citations_by_year["Year"], raw_citations_by_year.get("Scopus Citations", 0), marker="s", label="Scopus Citations")
        ax.plot(raw_citations_by_year["Year"], raw_citations_by_year.get("Total Citations Raw", 0), marker="^", label="Total Citations")
        ax.set_title("Annual evolution of citations (RAW data before deduplication)", weight="bold", pad=12)
        ax.set_xlabel("Year")
        ax.set_ylabel("Number of citations")
        ax.legend(loc="upper left")
        ax.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()

        if results_dir is not None:
            results_dir.mkdir(parents=True, exist_ok=True)
            saved_cites = results_dir / "raw_citations_by_year.png"
            fig.savefig(saved_cites, dpi=dpi, bbox_inches="tight")
            info("Gráfico guardado", f"RAW citas por año:\n{saved_cites}")

        if show:
            plt.show()
        else:
            plt.close(fig)

    return saved_counts, saved_cites
